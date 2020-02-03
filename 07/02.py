import json
import openpyxl
from openpyxl.utils import get_column_letter


def task_2(excel_path, json_path):
    wb = openpyxl.load_workbook(filename=excel_path)
    sheet = wb.active
    sheet.insert_cols(1, amount=1)
    students = {}
    average_marks = []
    for row_index in range(1, sheet.max_row + 1):
        name = sheet.cell(row_index, 2)
        surname = sheet.cell(row_index, 3)
        full_name = f'{name.value} {surname.value}'
        marks = []
        for col_index in range(4, sheet.max_column + 1):
            mark = sheet.cell(row_index, col_index).value
            if mark == 'end':
                break
            else:
                marks.append(mark)

        sheet.merge_cells(
            start_row=row_index, end_row=row_index,
            start_column=2, end_column=3
        )
        sheet.cell(row_index, 2).value = full_name
        max_column_letter = get_column_letter(3 + len(marks))
        sheet.cell(row_index, 1).value = (
            f'=AVERAGE(D{row_index}:{max_column_letter}{row_index})'
        )
        average_mark = sum(marks) / len(marks)
        students[full_name] = {
            'marks': marks,
            'average_mark': average_mark,
        }
        average_marks.append(average_mark)
    wb.save('table2.xlsx')

    result = {
        'students': students,
        'total_average_mark': sum(average_marks) / len(average_marks)
    }
    with open(json_path, 'w') as jf:
        jf.write(json.dumps(result))


if __name__ == '__main__':
    excel_path = 'table.xlsx'
    json_path = 'json2.json'
    task_2(excel_path, json_path)
